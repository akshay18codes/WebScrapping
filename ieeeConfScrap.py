from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
import os
from googlesearch import search
import urllib.parse

def scrape_ieee_conferences(url):
    # Chrome options
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Run in background
    chrome_options.add_argument("--disable-web-security")
    chrome_options.add_argument("--allow-running-insecure-content")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--start-maximized")
    
    # Setup WebDriver
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    
    
    conferences = []
    
    try:
        # Navigate to the page
        driver.get(url)
        
        # Wait and handle potential loading
        time.sleep(5)  # Initial wait
        
        # Try to find conference items with multiple strategies
        conference_strategies = [
            (By.CLASS_NAME, 'conference-item'),
            (By.XPATH, '//div[contains(@class, "conference-item")]'),
            (By.TAG_NAME, 'article')
        ]
        
        for strategy in conference_strategies:
            try:
                # Wait for elements with current strategy
                conference_items = WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located(strategy)
                )
                
                # Extract titles
                for item in conference_items:
                    
                        # Multiple ways to find title
                        title_strategies = [
                            lambda: item.find_element(By.CLASS_NAME, 'item-title').text,
                            lambda: item.find_element(By.TAG_NAME, 'h4').text,
                            lambda: item.text.split('\n')[0]
                        ]
                        
                        for title_strategy in title_strategies:
                            try:
                                title = title_strategy()
                                if title and title not in conferences:
                                    #print(f"Conference Title: {title}")
                                    conferences.append(title)
                            except Exception:
                                continue
                
                # If we found conferences, break the strategy loop
                if conferences:
                    break
            
            except Exception as strategy_error:
                print(f"Strategy {strategy} failed: {strategy_error}")
        
        return conferences
    
    except Exception as e:
        print(f"Comprehensive Error: {e}")
        return []
    
    finally:
        # Always close the browser
        
        driver.quit()
        return conferences

def append_to_excel(data_list, filename='output.xlsx', sheet_name='Sheet1'):

    # Check if the file exists
    if os.path.exists(filename):
        # Load existing workbook
        workbook = openpyxl.load_workbook(filename)
        
        # Check if sheet exists, if not create it
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        
        # Select the sheet
        sheet = workbook[sheet_name]
        
        # Find the first empty row
        last_row = sheet.max_row + 1
    else:
        # Create a new workbook if file doesn't exist
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        last_row = 1

    # Write the list contents starting from the first empty row
    for item in data_list:
        # Handle different types of list contents
        if isinstance(item, (list, tuple)):
            # If the item is a list or tuple, write each element in a separate column
            for col, value in enumerate(item, start=1):
                sheet.cell(row=last_row, column=col, value=value)
            last_row += 1
        else:
            # For simple lists, write each item in the first column
            sheet.cell(row=last_row, column=1, value=item)
            last_row += 1
    
    # Save the workbook
    workbook.save(filename)
    print(f"List contents appended to {filename}")

def read_excel_data(filename, sheet_name=None):
    
    # Load the workbook
    workbook = openpyxl.load_workbook(filename)
    
    # If no sheet name specified, use the active sheet
    if sheet_name is None:
        sheet = workbook.active
    else:
        sheet = workbook[sheet_name]
   
    # Method 2: Read specific column
    column_list = []
    for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
        column_list.append(row[0])
    
    return column_list

def perform_google_search(query, num_results=1):

    try:
        # URL encode the query
        encoded_query = urllib.parse.quote(query)
        # Perform the search
        search_results = []
        for result_url in search(query, num_results=num_results, advanced=True):
            search_results.append(result_url)
       
        if not search_results:
            print("No results found.")
        
        return result_url.url
    
    except Exception as e:
        print(f"An error occurred: {e}")
        return []

def main():
    urlpart1 = 'https://conferences.ieee.org/conferences_events/conferences/search?q=*&subsequent_q=&date=all&from=&to=&region=all&country=India&pos='
    urlpart2 = '&sortorder=desc&sponsor=&sponsor_type=all&state=all&field_of_interest=all&sortfield=relevance&searchmode=basic&virtualConfReadOnly=N&eventformat=hybrid'
    
    for i in range(4,7):
        url = "{}{}{}".format(urlpart1,i,urlpart2)
        
        conferences = scrape_ieee_conferences(url)
        # Example 1: Simple list of strings
        append_to_excel(conferences, 'conferences.xlsx','titles')
            # Read data from Excel
        try:
            # Read entire workbook
            excel_data = read_excel_data('conferences.xlsx')
            weblinks = []
            for query in excel_data:
                weblinks.append(perform_google_search(query))
            append_to_excel(weblinks, 'conferences.xlsx','links')

        except FileNotFoundError:
            print("Excel file not found!")
        except Exception as e:
            print(f"An error occurred: {e}")

if __name__ == "__main__":
    main()
